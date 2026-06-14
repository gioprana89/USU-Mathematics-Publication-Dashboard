import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIGURATION
# =====================================================
st.set_page_config(
    page_title="USU Mathematics Publication Dashboard",
    page_icon="📚",
    layout="wide"
)


# =====================================================
st.markdown(
    """
    Data Diperbarui pada 14 Juni 2026
  

    """
)
# =====================================================



st.title("📚 USU Mathematics Publication Dashboard")
st.caption(
    "Interactive dashboard for exploring journal and proceedings publication data of Mathematics Study Program lecturers "
    "at Universitas Sumatera Utara, with lecturer-year summaries and publication-style charts."
)


# =====================================================
# CONSTANTS
# =====================================================
REQUIRED_COLUMNS = [
    "Judul",
    "Author",
    "Nama Dosen USU",
    "Jurnal atau Prosiding",
    "Nama Jurnal atau Prosiding",
    "ISSN",
    "Tahun",
    "Volume",
    "Issue",
    "Halaman",
    "Scopus",
    "Sumber Artikel",
    "Sumber Artikel di Drive",
    "Scimago",
    "DOI",
    "Jumlah Sitasi Scopus",
    "Jumlah Sitasi Google Scholar",
]

DEFAULT_DISPLAY_COLUMNS = [
    "Judul",
    "Author",
    "Nama Dosen USU",
    "Jurnal atau Prosiding",
    "Nama Jurnal atau Prosiding",
    "Tahun",
    "Terindeks Scopus",
    "DOI",
    "Jumlah Sitasi Scopus",
    "Jumlah Sitasi Google Scholar",
    "Scopus",
    "Sumber Artikel",
    "Sumber Artikel di Drive",
    "Scimago",
]

LINK_COLUMNS = [
    "Scopus",
    "Sumber Artikel",
    "Sumber Artikel di Drive",
    "Scimago",
    "DOI",
]

COLOR_PALETTES = {
    "Scopus Orange Blue Green": [
        "#E97132", "#4472C4", "#70AD47", "#FFC000", "#7030A0",
        "#A5A5A5", "#C00000", "#00A6A6", "#595959", "#ED7D31"
    ],
    "Academic Blue": [
        "#1F4E79", "#5B9BD5", "#70AD47", "#FFC000", "#A5A5A5",
        "#ED7D31", "#7030A0", "#C00000", "#00A6A6", "#595959"
    ],
    "Nature Green": [
        "#375623", "#548235", "#70AD47", "#A9D18E", "#C5E0B4",
        "#806000", "#BF9000", "#FFC000", "#7F7F7F", "#595959"
    ],
    "Warm Publication": [
        "#C55A11", "#E97132", "#F4B183", "#FFD966", "#A64D79",
        "#7030A0", "#8064A2", "#B4A7D6", "#666666", "#999999"
    ],
    "Black Gray": [
        "#000000", "#404040", "#666666", "#808080", "#A6A6A6",
        "#BFBFBF", "#D9D9D9", "#595959", "#262626", "#8C8C8C"
    ],
    "High Contrast": [
        "#003F5C", "#BC5090", "#FFA600", "#58508D", "#FF6361",
        "#2F4B7C", "#665191", "#A05195", "#D45087", "#F95D6A"
    ]
}


# =====================================================
# HELPER FUNCTIONS: DATA LOADING AND CLEANING
# =====================================================
def read_uploaded_file(uploaded_file):
    file_name = uploaded_file.name.lower()

    if file_name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file)
        except UnicodeDecodeError:
            return pd.read_csv(uploaded_file, encoding="latin-1")

    if file_name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, engine="openpyxl")

    if file_name.endswith(".xls"):
        return pd.read_excel(uploaded_file)

    raise ValueError("Unsupported file format. Please upload CSV, XLSX, or XLS.")


@st.cache_data
def read_default_dataset(path_string):
    return pd.read_excel(path_string, engine="openpyxl")


def find_default_dataset():
    candidates = [
        Path("data publikasi new 2.xlsx"),
        Path(__file__).parent / "data publikasi new 2.xlsx" if "__file__" in globals() else Path("data publikasi new 2.xlsx"),
    ]

    for path in candidates:
        if path.exists():
            return path

    return None


def clean_dataframe(df):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    df = df.dropna(how="all")

    unnamed_cols = [
        col for col in df.columns
        if str(col).lower().startswith("unnamed") and df[col].isna().all()
    ]
    if unnamed_cols:
        df = df.drop(columns=unnamed_cols)

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).replace("nan", np.nan)
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    if "Tahun" in df.columns:
        year_numeric = pd.to_numeric(df["Tahun"], errors="coerce")
        if year_numeric.notna().any():
            df["Tahun"] = year_numeric.astype("Int64").astype(str).replace("<NA>", np.nan)
        else:
            df["Tahun"] = df["Tahun"].astype(str).str.strip()

    for citation_col in ["Jumlah Sitasi Scopus", "Jumlah Sitasi Google Scholar"]:
        if citation_col in df.columns:
            df[citation_col] = pd.to_numeric(df[citation_col], errors="coerce")

    if "Terindeks Scopus" not in df.columns:
        if "Scopus" in df.columns:
            df["Terindeks Scopus"] = np.where(
                df["Scopus"].notna() & (df["Scopus"].astype(str).str.strip() != ""),
                "Ya",
                "Tidak"
            )
        else:
            df["Terindeks Scopus"] = "Tidak tersedia"

    df["Terindeks Scopus"] = (
        df["Terindeks Scopus"]
        .fillna("Tidak")
        .astype(str)
        .str.strip()
        .replace({"": "Tidak"})
    )

    return df


def split_lecturer_names(value):
    if pd.isna(value):
        return []

    text = str(value).strip()
    if not text:
        return []

    # Dataset uses semicolon-separated lecturer names.
    parts = [part.strip() for part in text.replace("\n", ";").split(";")]
    return [part for part in parts if part]


def add_publication_id(df):
    df = df.copy()
    df["Publication ID"] = np.arange(1, len(df) + 1)
    return df


def make_lecturer_long(df):
    records = []

    if "Nama Dosen USU" not in df.columns:
        return pd.DataFrame(columns=["Publication ID", "Nama Dosen USU (Individual)"])

    for _, row in df.iterrows():
        lecturer_names = split_lecturer_names(row.get("Nama Dosen USU"))
        for lecturer in lecturer_names:
            records.append({
                "Publication ID": row["Publication ID"],
                "Nama Dosen USU (Individual)": lecturer,
            })

    return pd.DataFrame(records)


def safe_unique(series):
    values = series.dropna().astype(str).str.strip()
    values = values[values != ""]
    return sorted(values.unique().tolist(), key=lambda x: x.lower())


def safe_year_sort(values):
    values = list(values)
    numeric = pd.to_numeric(pd.Series(values), errors="coerce")
    if numeric.notna().all():
        return [x for x, _ in sorted(zip(values, numeric), key=lambda item: item[1])]
    return sorted(values, key=lambda x: str(x).lower())


def apply_filters(df, lecturer_long, selected_lecturers, selected_pub_types, selected_years, selected_scopus_status):
    filtered = df.copy()

    if selected_lecturers and "Nama Dosen USU" in filtered.columns:
        matched_ids = lecturer_long[
            lecturer_long["Nama Dosen USU (Individual)"].isin(selected_lecturers)
        ]["Publication ID"].unique()
        filtered = filtered[filtered["Publication ID"].isin(matched_ids)]

    if selected_pub_types and "Jurnal atau Prosiding" in filtered.columns:
        filtered = filtered[filtered["Jurnal atau Prosiding"].astype(str).isin(selected_pub_types)]

    if selected_years and "Tahun" in filtered.columns:
        filtered = filtered[filtered["Tahun"].astype(str).isin([str(y) for y in selected_years])]

    if selected_scopus_status and "Terindeks Scopus" in filtered.columns:
        filtered = filtered[filtered["Terindeks Scopus"].astype(str).isin(selected_scopus_status)]

    return filtered


# =====================================================
# HELPER FUNCTIONS: TABLES
# =====================================================
def build_selected_information_table(filtered_df, selected_info_cols):
    if not selected_info_cols:
        return pd.DataFrame()

    table = (
        filtered_df
        .groupby(selected_info_cols, dropna=False)
        .size()
        .reset_index(name="Jumlah Publikasi")
    )

    total = table["Jumlah Publikasi"].sum()
    table["Persentase (%)"] = np.where(
        total > 0,
        table["Jumlah Publikasi"] / total * 100,
        0
    )
    table["Persentase (%)"] = table["Persentase (%)"].round(2)

    return table.sort_values("Jumlah Publikasi", ascending=False)


def build_summary_tables(filtered_df, lecturer_long):
    summaries = {}

    if filtered_df.empty:
        empty = pd.DataFrame()
        return {
            "summary_by_lecturer": empty,
            "summary_by_year": empty,
            "summary_by_type": empty,
            "summary_by_scopus": empty,
            "cross_year_type": empty,
            "cross_year_scopus": empty,
        }

    filtered_ids = filtered_df["Publication ID"].unique()
    lecturer_filtered = lecturer_long[lecturer_long["Publication ID"].isin(filtered_ids)].copy()

    if not lecturer_filtered.empty:
        lecturer_summary = (
            lecturer_filtered
            .groupby("Nama Dosen USU (Individual)")
            .agg(**{"Jumlah Publikasi": ("Publication ID", "nunique")})
            .reset_index()
            .sort_values("Jumlah Publikasi", ascending=False)
        )
    else:
        lecturer_summary = pd.DataFrame(columns=["Nama Dosen USU (Individual)", "Jumlah Publikasi"])

    summaries["summary_by_lecturer"] = lecturer_summary

    if "Tahun" in filtered_df.columns:
        summaries["summary_by_year"] = (
            filtered_df
            .groupby("Tahun", dropna=False)
            .agg(**{"Jumlah Publikasi": ("Publication ID", "nunique")})
            .reset_index()
            .sort_values("Tahun", key=lambda s: pd.to_numeric(s, errors="coerce"))
        )
    else:
        summaries["summary_by_year"] = pd.DataFrame()

    if "Jurnal atau Prosiding" in filtered_df.columns:
        summaries["summary_by_type"] = (
            filtered_df
            .groupby("Jurnal atau Prosiding", dropna=False)
            .agg(**{"Jumlah Publikasi": ("Publication ID", "nunique")})
            .reset_index()
            .sort_values("Jumlah Publikasi", ascending=False)
        )
    else:
        summaries["summary_by_type"] = pd.DataFrame()

    if "Terindeks Scopus" in filtered_df.columns:
        summaries["summary_by_scopus"] = (
            filtered_df
            .groupby("Terindeks Scopus", dropna=False)
            .agg(**{"Jumlah Publikasi": ("Publication ID", "nunique")})
            .reset_index()
            .sort_values("Jumlah Publikasi", ascending=False)
        )
    else:
        summaries["summary_by_scopus"] = pd.DataFrame()

    if "Tahun" in filtered_df.columns and "Jurnal atau Prosiding" in filtered_df.columns:
        summaries["cross_year_type"] = pd.crosstab(
            filtered_df["Tahun"],
            filtered_df["Jurnal atau Prosiding"],
            margins=True,
            margins_name="Total"
        )
    else:
        summaries["cross_year_type"] = pd.DataFrame()

    if "Tahun" in filtered_df.columns and "Terindeks Scopus" in filtered_df.columns:
        summaries["cross_year_scopus"] = pd.crosstab(
            filtered_df["Tahun"],
            filtered_df["Terindeks Scopus"],
            margins=True,
            margins_name="Total"
        )
    else:
        summaries["cross_year_scopus"] = pd.DataFrame()

    return summaries


def build_lecturer_publication_long(filtered_df, lecturer_long):
    """Create lecturer-publication long data after dashboard filters are applied."""
    required_base_cols = ["Publication ID"]
    optional_cols = [
        "Judul",
        "Nama Dosen USU",
        "Jurnal atau Prosiding",
        "Tahun",
        "Terindeks Scopus",
        "Nama Jurnal atau Prosiding",
        "Jumlah Sitasi Scopus",
        "Jumlah Sitasi Google Scholar",
    ]

    if filtered_df.empty or lecturer_long.empty or "Publication ID" not in filtered_df.columns:
        return pd.DataFrame()

    available_cols = required_base_cols + [col for col in optional_cols if col in filtered_df.columns]
    filtered_ids = filtered_df["Publication ID"].unique()

    source = (
        lecturer_long[lecturer_long["Publication ID"].isin(filtered_ids)]
        .merge(filtered_df[available_cols], on="Publication ID", how="left")
    )

    if source.empty:
        return source

    for col in ["Nama Dosen USU (Individual)", "Jurnal atau Prosiding", "Tahun", "Terindeks Scopus"]:
        if col in source.columns:
            source[col] = source[col].fillna("Tidak tersedia").astype(str).str.strip()
            source[col] = source[col].replace({"": "Tidak tersedia", "nan": "Tidak tersedia"})

    if "Tahun" in source.columns:
        source["Tahun"] = source["Tahun"].astype(str)

    # Prevent accidental duplicate lecturer-publication rows.
    source = source.drop_duplicates(subset=["Publication ID", "Nama Dosen USU (Individual)"])
    return source


def build_lecturer_year_type_summary(filtered_df, lecturer_long):
    """Build summary tables: lecturer x year x publication type, lecturer total, and year total."""
    long_df = build_lecturer_publication_long(filtered_df, lecturer_long)

    empty = pd.DataFrame()
    if long_df.empty:
        return empty, empty, empty, empty

    required_cols = ["Nama Dosen USU (Individual)", "Tahun", "Jurnal atau Prosiding"]
    if any(col not in long_df.columns for col in required_cols):
        return long_df, empty, empty, empty

    summary = (
        long_df
        .groupby(required_cols, dropna=False)["Publication ID"]
        .nunique()
        .unstack(fill_value=0)
        .reset_index()
    )

    type_cols = [col for col in summary.columns if col not in ["Nama Dosen USU (Individual)", "Tahun"]]
    summary["Total Publikasi"] = summary[type_cols].sum(axis=1) if type_cols else 0

    summary = summary.sort_values(
        ["Nama Dosen USU (Individual)", "Tahun"],
        key=lambda s: pd.to_numeric(s, errors="coerce") if s.name == "Tahun" else s.astype(str).str.lower()
    )

    lecturer_total = (
        long_df
        .groupby(["Nama Dosen USU (Individual)", "Jurnal atau Prosiding"], dropna=False)["Publication ID"]
        .nunique()
        .unstack(fill_value=0)
        .reset_index()
    )
    total_type_cols = [col for col in lecturer_total.columns if col != "Nama Dosen USU (Individual)"]
    lecturer_total["Total Publikasi"] = lecturer_total[total_type_cols].sum(axis=1) if total_type_cols else 0
    lecturer_total = lecturer_total.sort_values("Total Publikasi", ascending=False)

    year_total = (
        long_df
        .groupby(["Tahun", "Jurnal atau Prosiding"], dropna=False)["Publication ID"]
        .nunique()
        .unstack(fill_value=0)
        .reset_index()
    )
    year_type_cols = [col for col in year_total.columns if col != "Tahun"]
    year_total["Total Publikasi"] = year_total[year_type_cols].sum(axis=1) if year_type_cols else 0
    year_total = year_total.sort_values("Tahun", key=lambda s: pd.to_numeric(s, errors="coerce"))

    return long_df, summary, lecturer_total, year_total


def build_lecturer_year_chart_table(
    lecturer_publication_long,
    x_dimension,
    stack_dimension,
    value_mode,
    percentage_basis
):
    """Build chart table for lecturer-year publication summary visualization."""
    if lecturer_publication_long is None or lecturer_publication_long.empty:
        return pd.DataFrame()

    source = lecturer_publication_long.copy()

    if "Tahun" in source.columns and "Nama Dosen USU (Individual)" in source.columns:
        source["Tahun | Dosen"] = (
            source["Tahun"].astype(str)
            + " | "
            + source["Nama Dosen USU (Individual)"].astype(str)
        )

    if x_dimension not in source.columns:
        return pd.DataFrame()

    source = source.dropna(subset=[x_dimension]).copy()
    source[x_dimension] = source[x_dimension].astype(str).str.strip()
    source = source[source[x_dimension] != ""]

    # Count one publication once for each individual lecturer.
    duplicate_subset = ["Publication ID", "Nama Dosen USU (Individual)", x_dimension]
    if stack_dimension != "Tidak ada / single bar" and stack_dimension in source.columns:
        duplicate_subset.append(stack_dimension)
    source = source.drop_duplicates(subset=[col for col in duplicate_subset if col in source.columns])

    if stack_dimension == "Tidak ada / single bar":
        count_table = (
            source
            .groupby(x_dimension, dropna=False)
            .size()
            .to_frame("Jumlah Publikasi")
        )
    else:
        if stack_dimension not in source.columns:
            return pd.DataFrame()
        source = source.dropna(subset=[stack_dimension]).copy()
        source[stack_dimension] = source[stack_dimension].astype(str).str.strip()
        source = source[source[stack_dimension] != ""]

        count_table = pd.crosstab(
            source[x_dimension],
            source[stack_dimension]
        )

    plot_table = count_table.astype(float)

    if value_mode == "Persentase (%)":
        if percentage_basis == "Setiap bar menjadi 100%" and stack_dimension != "Tidak ada / single bar":
            denominator = plot_table.sum(axis=1).replace(0, np.nan)
            plot_table = plot_table.div(denominator, axis=0) * 100
        else:
            total = plot_table.values.sum()
            plot_table = plot_table / total * 100 if total else plot_table * 0

    return plot_table.fillna(0)


def format_percentage_table(count_table, basis="row"):
    if count_table.empty:
        return count_table

    data = count_table.copy()
    if "Total" in data.index:
        data = data.drop(index="Total")
    if "Total" in data.columns:
        data = data.drop(columns="Total")

    if basis == "row":
        denominator = data.sum(axis=1).replace(0, np.nan)
        pct = data.div(denominator, axis=0) * 100
    elif basis == "column":
        denominator = data.sum(axis=0).replace(0, np.nan)
        pct = data.div(denominator, axis=1) * 100
    else:
        total = data.values.sum()
        pct = data / total * 100 if total else data * 0

    return pct.round(2)


# =====================================================
# HELPER FUNCTIONS: CHARTS
# =====================================================
def reorder_axis(table, axis_order_option, custom_order_text, axis=0):
    if table.empty:
        return table

    table = table.copy()
    labels = list(table.index if axis == 0 else table.columns)

    if axis_order_option == "Naik (A-Z / kecil-besar)":
        ordered = safe_year_sort(labels)
    elif axis_order_option == "Turun (Z-A / besar-kecil)":
        ordered = safe_year_sort(labels)[::-1]
    elif axis_order_option == "Total tertinggi":
        totals = table.sum(axis=1 if axis == 0 else 0)
        ordered = totals.sort_values(ascending=False).index.tolist()
    elif axis_order_option == "Total terendah":
        totals = table.sum(axis=1 if axis == 0 else 0)
        ordered = totals.sort_values(ascending=True).index.tolist()
    elif axis_order_option == "Custom":
        typed = [item.strip() for item in custom_order_text.split(",") if item.strip()]
        ordered = [item for item in typed if item in labels]
        ordered += [item for item in labels if item not in ordered]
    else:
        ordered = labels

    if axis == 0:
        return table.reindex(ordered)

    return table.reindex(columns=ordered)


def make_chart_source(filtered_df, lecturer_long, x_dimension, stack_dimension):
    lecturer_label = "Nama Dosen USU (Individual)"
    needs_lecturer = x_dimension == lecturer_label or stack_dimension == lecturer_label

    if needs_lecturer:
        filtered_ids = filtered_df["Publication ID"].unique()
        base = lecturer_long[lecturer_long["Publication ID"].isin(filtered_ids)].copy()
        base = base.merge(filtered_df, on="Publication ID", how="left", suffixes=("", "_publication"))
    else:
        base = filtered_df.copy()

    return base


def build_chart_table(
    filtered_df,
    lecturer_long,
    x_dimension,
    stack_dimension,
    value_mode,
    percentage_basis
):
    source = make_chart_source(filtered_df, lecturer_long, x_dimension, stack_dimension)

    if source.empty or x_dimension not in source.columns:
        return pd.DataFrame()

    source = source.dropna(subset=[x_dimension]).copy()
    source[x_dimension] = source[x_dimension].astype(str).str.strip()
    source = source[source[x_dimension] != ""]

    if stack_dimension == "Tidak ada / single bar":
        count_table = (
            source
            .groupby(x_dimension, dropna=False)
            .agg(**{"Jumlah Publikasi": ("Publication ID", "nunique")})
        )
        plot_table = count_table.astype(float)
    else:
        if stack_dimension not in source.columns:
            return pd.DataFrame()

        source = source.dropna(subset=[stack_dimension]).copy()
        source[stack_dimension] = source[stack_dimension].astype(str).str.strip()
        source = source[source[stack_dimension] != ""]

        # Publication ID is used to avoid duplicate records when the source is publication-level.
        count_table = pd.crosstab(
            source[x_dimension],
            source[stack_dimension]
        )
        plot_table = count_table.astype(float)

    if value_mode == "Persentase (%)":
        if percentage_basis == "Setiap bar menjadi 100%" and stack_dimension != "Tidak ada / single bar":
            denominator = plot_table.sum(axis=1).replace(0, np.nan)
            plot_table = plot_table.div(denominator, axis=0) * 100
        else:
            total = plot_table.values.sum()
            plot_table = plot_table / total * 100 if total else plot_table * 0

    return plot_table.fillna(0)


def create_publication_bar_chart(
    plot_table,
    value_mode,
    chart_type,
    category_colors,
    single_bar_color,
    title,
    subtitle,
    x_label,
    y_label,
    legend_title,
    figure_width,
    figure_height,
    bar_width,
    show_grid,
    show_value_labels,
    label_decimal_digits,
    label_min_value,
    x_tick_rotation,
    font_family,
    title_font_size,
    subtitle_font_size,
    axis_font_size,
    tick_font_size,
    legend_font_size,
    label_font_size,
    title_color,
    subtitle_color,
    axis_color,
    tick_color,
    legend_color,
    label_color,
    edge_color,
    edge_width,
    legend_position
):
    plt.rcParams["font.family"] = font_family

    fig, ax = plt.subplots(figsize=(figure_width, figure_height))

    x_labels = plot_table.index.astype(str).tolist()
    x = np.arange(len(x_labels))
    categories = plot_table.columns.tolist()

    is_single = categories == ["Jumlah Publikasi"] or len(categories) == 1 and categories[0] == "Jumlah Publikasi"

    if is_single:
        values = plot_table.iloc[:, 0].fillna(0).to_numpy(dtype=float)
        bars = ax.bar(
            x,
            values,
            width=bar_width,
            color=single_bar_color,
            edgecolor=edge_color,
            linewidth=edge_width
        )

        if show_value_labels:
            for bar, value in zip(bars, values):
                if value <= 0 or value < label_min_value:
                    continue
                label = f"{value:.{label_decimal_digits}f}%" if value_mode == "Persentase (%)" else f"{int(round(value))}"
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    label,
                    ha="center",
                    va="bottom",
                    fontsize=label_font_size,
                    color=label_color
                )
    elif chart_type == "Grouped Bar":
        n_categories = len(categories)
        total_width = bar_width
        single_width = total_width / max(n_categories, 1)

        for idx, category in enumerate(categories):
            values = plot_table[category].fillna(0).to_numpy(dtype=float)
            color = category_colors.get(category, "#808080")
            offset = (idx - (n_categories - 1) / 2) * single_width

            bars = ax.bar(
                x + offset,
                values,
                width=single_width * 0.92,
                label=category,
                color=color,
                edgecolor=edge_color,
                linewidth=edge_width
            )

            if show_value_labels:
                for bar, value in zip(bars, values):
                    if value <= 0 or value < label_min_value:
                        continue
                    label = f"{value:.{label_decimal_digits}f}%" if value_mode == "Persentase (%)" else f"{int(round(value))}"
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height(),
                        label,
                        ha="center",
                        va="bottom",
                        fontsize=label_font_size,
                        color=label_color
                    )
    else:
        bottom = np.zeros(len(plot_table))
        for category in categories:
            values = plot_table[category].fillna(0).to_numpy(dtype=float)
            color = category_colors.get(category, "#808080")

            ax.bar(
                x,
                values,
                bottom=bottom,
                width=bar_width,
                label=category,
                color=color,
                edgecolor=edge_color,
                linewidth=edge_width
            )

            if show_value_labels:
                for i, value in enumerate(values):
                    if value <= 0 or value < label_min_value:
                        continue
                    label = f"{value:.{label_decimal_digits}f}%" if value_mode == "Persentase (%)" else f"{int(round(value))}"
                    ax.text(
                        x[i],
                        bottom[i] + value / 2,
                        label,
                        ha="center",
                        va="center",
                        fontsize=label_font_size,
                        color=label_color
                    )
            bottom += values

    ax.set_title(title, fontsize=title_font_size, fontweight="bold", color=title_color, pad=18)

    if subtitle.strip():
        ax.text(
            0.5,
            1.01,
            subtitle,
            transform=ax.transAxes,
            ha="center",
            va="bottom",
            fontsize=subtitle_font_size,
            color=subtitle_color
        )

    ax.set_xlabel(x_label, fontsize=axis_font_size, color=axis_color, labelpad=10)
    ax.set_ylabel(y_label, fontsize=axis_font_size, color=axis_color, labelpad=10)

    ax.set_xticks(x)
    ax.set_xticklabels(
        x_labels,
        rotation=x_tick_rotation,
        ha="right" if x_tick_rotation > 0 else "center",
        fontsize=tick_font_size,
        color=tick_color
    )
    ax.tick_params(axis="y", labelsize=tick_font_size, colors=tick_color)

    if value_mode == "Persentase (%)":
        if chart_type == "Stacked Bar" and not is_single:
            ax.set_ylim(0, 100)

    if show_grid:
        ax.grid(axis="y", linestyle="--", linewidth=0.6, alpha=0.45)
        ax.set_axisbelow(True)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    if not is_single:
        if legend_position == "Di luar kanan":
            legend = ax.legend(
                title=legend_title,
                loc="upper left",
                bbox_to_anchor=(1.02, 1),
                frameon=True,
                fontsize=legend_font_size,
                title_fontsize=legend_font_size
            )
        elif legend_position == "Bawah":
            legend = ax.legend(
                title=legend_title,
                loc="upper center",
                bbox_to_anchor=(0.5, -0.20),
                ncol=min(4, max(1, len(categories))),
                frameon=True,
                fontsize=legend_font_size,
                title_fontsize=legend_font_size
            )
        else:
            legend = ax.legend(
                title=legend_title,
                loc="upper right",
                frameon=True,
                fontsize=legend_font_size,
                title_fontsize=legend_font_size
            )

        for text in legend.get_texts():
            text.set_color(legend_color)
        legend.get_title().set_color(legend_color)

    fig.tight_layout()
    return fig


def fig_to_png_bytes(fig, dpi, transparent_background=False):
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="none" if transparent_background else "white",
        transparent=transparent_background
    )
    buffer.seek(0)
    return buffer


# =====================================================
# HELPER FUNCTIONS: EXCEL EXPORT
# =====================================================
def auto_style_workbook(workbook):
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF" if sheet_name == "Info" else "000000")
                cell.fill = title_fill if sheet_name == "Info" else header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)


def dataframe_to_excel_bytes(
    filtered_df,
    selected_info_table,
    summaries,
    chart_table,
    lecturer_publication_long,
    lecturer_year_type_summary,
    lecturer_overall_type_summary,
    lecturer_year_total_type_summary,
    lecturer_year_chart_table,
    selected_lecturers,
    selected_pub_types,
    selected_years,
    selected_scopus_status
):
    output = BytesIO()

    info_df = pd.DataFrame({
        "Item": [
            "Selected Lecturers",
            "Selected Publication Types",
            "Selected Years",
            "Selected Scopus Status",
            "Number of Filtered Publications",
        ],
        "Value": [
            "; ".join(selected_lecturers) if selected_lecturers else "All",
            "; ".join(selected_pub_types) if selected_pub_types else "All",
            "; ".join([str(x) for x in selected_years]) if selected_years else "All",
            "; ".join(selected_scopus_status) if selected_scopus_status else "All",
            len(filtered_df),
        ]
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        info_df.to_excel(writer, sheet_name="Info", index=False)
        filtered_df.to_excel(writer, sheet_name="Filtered Data", index=False)

        if selected_info_table is not None and not selected_info_table.empty:
            selected_info_table.to_excel(writer, sheet_name="Selected Info Summary", index=False)

        for key, table in summaries.items():
            sheet_name = key.replace("summary_by_", "By ").replace("cross_", "Cross ")[:31]
            if table is not None and not table.empty:
                table.to_excel(writer, sheet_name=sheet_name, index=True if isinstance(table.index, pd.Index) else False)

        if lecturer_publication_long is not None and not lecturer_publication_long.empty:
            lecturer_publication_long.to_excel(writer, sheet_name="Lecturer Publication Long", index=False)

        if lecturer_year_type_summary is not None and not lecturer_year_type_summary.empty:
            lecturer_year_type_summary.to_excel(writer, sheet_name="Lecturer Year Type", index=False)

        if lecturer_overall_type_summary is not None and not lecturer_overall_type_summary.empty:
            lecturer_overall_type_summary.to_excel(writer, sheet_name="Lecturer Type Total", index=False)

        if lecturer_year_total_type_summary is not None and not lecturer_year_total_type_summary.empty:
            lecturer_year_total_type_summary.to_excel(writer, sheet_name="Year Type Total", index=False)

        if lecturer_year_chart_table is not None and not lecturer_year_chart_table.empty:
            lecturer_year_chart_table.to_excel(writer, sheet_name="Lecturer Year Chart")

        if chart_table is not None and not chart_table.empty:
            chart_table.to_excel(writer, sheet_name="Chart Table")

        auto_style_workbook(writer.book)

    output.seek(0)
    return output.getvalue()


def display_dataframe_with_links(dataframe, use_container_width=True):
    column_config = {}
    for col in dataframe.columns:
        if col in LINK_COLUMNS:
            column_config[col] = st.column_config.LinkColumn(col)

    st.dataframe(
        dataframe,
        use_container_width=use_container_width,
        column_config=column_config if column_config else None,
        hide_index=True
    )


# =====================================================
# LOAD DATA
# =====================================================
st.sidebar.header("1. Data Source")

uploaded_file = st.sidebar.file_uploader(
    "Upload file Excel atau CSV",
    type=["xlsx", "xls", "csv"]
)

use_default_data = st.sidebar.checkbox(
    "Gunakan data publikasi.xlsx jika tersedia",
    value=True
)

try:
    if uploaded_file is not None:
        df_raw = read_uploaded_file(uploaded_file)
        active_source_name = uploaded_file.name
    elif use_default_data:
        default_path = find_default_dataset()
        if default_path is None:
            st.info("File default 'data publikasi.xlsx' belum ditemukan. Silakan upload file terlebih dahulu.")
            st.stop()
        df_raw = read_default_dataset(str(default_path))
        active_source_name = default_path.name
    else:
        st.info("Silakan upload dataset Excel/CSV terlebih dahulu.")
        st.stop()
except Exception as e:
    st.error(f"Gagal membaca dataset: {e}")
    st.info("Jika membaca file Excel .xlsx gagal, jalankan: python -m pip install openpyxl")
    st.stop()


df = clean_dataframe(df_raw)
df = add_publication_id(df)
lecturer_long = make_lecturer_long(df)

if df.empty:
    st.error("Dataset kosong.")
    st.stop()

columns = df.columns.tolist()
missing_columns = [col for col in REQUIRED_COLUMNS if col not in columns]

if missing_columns:
    st.warning(
        "Beberapa kolom standar tidak ditemukan: " + ", ".join(missing_columns) +
        ". Dashboard tetap berjalan menggunakan kolom yang tersedia."
    )


# =====================================================
# SIDEBAR FILTERS
# =====================================================
st.sidebar.header("2. Pilih Nama Dosen")

lecturer_options = (
    lecturer_long["Nama Dosen USU (Individual)"].dropna().unique().tolist()
    if not lecturer_long.empty else []
)
lecturer_options = sorted(lecturer_options, key=lambda x: x.lower())

selected_lecturers = st.sidebar.multiselect(
    "Silahkan pilih nama dosen yang datanya ingin anda akses:",
    lecturer_options,
    default=lecturer_options
)

st.sidebar.header("3. Filter Informasi Publikasi")

pub_type_options = safe_unique(df["Jurnal atau Prosiding"]) if "Jurnal atau Prosiding" in df.columns else []
selected_pub_types = st.sidebar.multiselect(
    "Jurnal atau Prosiding",
    pub_type_options,
    default=pub_type_options
)

year_options = safe_year_sort(safe_unique(df["Tahun"])) if "Tahun" in df.columns else []
selected_years = st.sidebar.multiselect(
    "Tahun",
    year_options,
    default=year_options
)

scopus_options = safe_unique(df["Terindeks Scopus"]) if "Terindeks Scopus" in df.columns else []
selected_scopus_status = st.sidebar.multiselect(
    "Terindeks Scopus",
    scopus_options,
    default=scopus_options
)

info_filter_columns_available = [
    col for col in ["Jurnal atau Prosiding", "Tahun", "Terindeks Scopus"]
    if col in df.columns
]

selected_info_cols = st.sidebar.multiselect(
    "Pilih kolom informasi untuk tabel rekap",
    info_filter_columns_available,
    default=info_filter_columns_available
)

available_display_columns = [col for col in DEFAULT_DISPLAY_COLUMNS if col in df.columns]
if not available_display_columns:
    available_display_columns = [col for col in df.columns if col != "Publication ID"][:8]

display_columns = st.sidebar.multiselect(
    "Pilih kolom yang ditampilkan pada tabel detail",
    [col for col in df.columns if col != "Publication ID"],
    default=available_display_columns
)

filtered_df = apply_filters(
    df=df,
    lecturer_long=lecturer_long,
    selected_lecturers=selected_lecturers,
    selected_pub_types=selected_pub_types,
    selected_years=selected_years,
    selected_scopus_status=selected_scopus_status
)

if filtered_df.empty:
    st.warning("Tidak ada data yang sesuai dengan filter yang dipilih.")
    st.stop()

selected_info_table = build_selected_information_table(filtered_df, selected_info_cols)
summaries = build_summary_tables(filtered_df, lecturer_long)

(
    lecturer_publication_long,
    lecturer_year_type_summary,
    lecturer_overall_type_summary,
    lecturer_year_total_type_summary,
) = build_lecturer_year_type_summary(filtered_df, lecturer_long)


# =====================================================
# SIDEBAR CHART SETTINGS
# =====================================================
st.sidebar.header("4. Pengaturan Visualisasi")

lecturer_chart_label = "Nama Dosen USU (Individual)"
chart_dimension_options = [lecturer_chart_label]
for col in ["Tahun", "Jurnal atau Prosiding", "Terindeks Scopus"]:
    if col in df.columns:
        chart_dimension_options.append(col)

x_dimension = st.sidebar.selectbox(
    "Variabel sumbu X grafik",
    chart_dimension_options,
    index=chart_dimension_options.index("Tahun") if "Tahun" in chart_dimension_options else 0
)

stack_options = ["Tidak ada / single bar"] + [col for col in chart_dimension_options if col != x_dimension]
stack_dimension = st.sidebar.selectbox(
    "Variabel stacked / legenda grafik",
    stack_options,
    index=0 if "Jurnal atau Prosiding" not in stack_options else stack_options.index("Jurnal atau Prosiding")
)

value_mode = st.sidebar.radio(
    "Tampilkan nilai sebagai",
    ["Frekuensi", "Persentase (%)"],
    horizontal=True
)

percentage_basis = st.sidebar.selectbox(
    "Basis persentase",
    ["Setiap bar menjadi 100%", "Dari total seluruh data"],
    index=0,
    disabled=value_mode == "Frekuensi" or stack_dimension == "Tidak ada / single bar"
)

chart_type = st.sidebar.radio(
    "Tipe grafik",
    ["Stacked Bar", "Grouped Bar"],
    horizontal=True,
    disabled=stack_dimension == "Tidak ada / single bar"
)

order_options = [
    "Sesuai data",
    "Naik (A-Z / kecil-besar)",
    "Turun (Z-A / besar-kecil)",
    "Total tertinggi",
    "Total terendah",
    "Custom"
]

x_order_option = st.sidebar.selectbox("Urutan label sumbu X", order_options, index=1)
x_custom_order = ""
if x_order_option == "Custom":
    x_custom_order = st.sidebar.text_area(
        "Urutan custom sumbu X, pisahkan dengan koma",
        value="",
        height=80
    )

stack_order_option = "Sesuai data"
stack_custom_order = ""
if stack_dimension != "Tidak ada / single bar":
    stack_order_option = st.sidebar.selectbox("Urutan kategori legenda/stacked", order_options, index=0)
    if stack_order_option == "Custom":
        stack_custom_order = st.sidebar.text_area(
            "Urutan custom kategori stacked, pisahkan dengan koma",
            value="",
            height=80
        )

raw_chart_table = build_chart_table(
    filtered_df=filtered_df,
    lecturer_long=lecturer_long,
    x_dimension=x_dimension,
    stack_dimension=stack_dimension,
    value_mode=value_mode,
    percentage_basis=percentage_basis
)

chart_table = reorder_axis(raw_chart_table, x_order_option, x_custom_order, axis=0)
if stack_dimension != "Tidak ada / single bar":
    chart_table = reorder_axis(chart_table, stack_order_option, stack_custom_order, axis=1)

if chart_table.empty:
    st.warning("Tabel grafik kosong. Coba ubah filter atau variabel grafik.")
    st.stop()

st.sidebar.header("5. Warna Grafik")

palette_name = st.sidebar.selectbox("Pilih palet warna", list(COLOR_PALETTES.keys()), index=0)
palette = COLOR_PALETTES[palette_name]

single_bar_color = st.sidebar.color_picker("Warna single bar", palette[0])

category_colors = {}
for i, category in enumerate(chart_table.columns):
    if category == "Jumlah Publikasi":
        continue
    category_colors[category] = st.sidebar.color_picker(
        f"Warna: {category}",
        palette[i % len(palette)]
    )

st.sidebar.header("6. Teks, Layout, dan Ekspor PNG")

font_family = st.sidebar.selectbox(
    "Jenis font",
    ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"],
    index=0
)

default_chart_title = f"Publication Distribution by {x_dimension}"
if stack_dimension != "Tidak ada / single bar":
    default_chart_title += f" and {stack_dimension}"

chart_title = st.sidebar.text_area("Judul grafik", value=default_chart_title, height=70)
chart_subtitle = st.sidebar.text_input("Subjudul", value="Based on selected lecturer publication data")
x_label = st.sidebar.text_input("Label sumbu X", value=x_dimension)
y_label = st.sidebar.text_input("Label sumbu Y", value="Frequency" if value_mode == "Frekuensi" else "Percentage (%)")
legend_title = st.sidebar.text_input("Judul legenda", value=stack_dimension if stack_dimension != "Tidak ada / single bar" else "")

figure_width = st.sidebar.slider("Lebar grafik", 6.0, 28.0, 13.0, 0.5)
figure_height = st.sidebar.slider("Tinggi grafik", 4.0, 20.0, 7.5, 0.5)
bar_width = st.sidebar.slider("Lebar batang", 0.20, 1.00, 0.72, 0.02)
x_tick_rotation = st.sidebar.slider("Rotasi teks sumbu X", 0, 90, 25, 5)

show_grid = st.sidebar.checkbox("Tampilkan grid horizontal", value=True)
show_value_labels = st.sidebar.checkbox("Tampilkan label nilai", value=True)
label_decimal_digits = st.sidebar.slider("Digit di belakang koma untuk label", 0, 4, 2)
label_min_value = st.sidebar.number_input("Minimal nilai yang diberi label", min_value=0.0, value=1.0, step=1.0)

edge_color = st.sidebar.color_picker("Warna garis tepi batang", "#FFFFFF")
edge_width = st.sidebar.slider("Ketebalan garis tepi batang", 0.0, 3.0, 0.6, 0.1)
legend_position = st.sidebar.selectbox("Posisi legenda", ["Di luar kanan", "Di dalam kanan atas", "Bawah"], index=0)

title_font_size = st.sidebar.slider("Ukuran font judul", 10, 40, 18)
subtitle_font_size = st.sidebar.slider("Ukuran font subjudul", 8, 30, 12)
axis_font_size = st.sidebar.slider("Ukuran font label sumbu", 8, 28, 12)
tick_font_size = st.sidebar.slider("Ukuran font tick", 6, 24, 10)
legend_font_size = st.sidebar.slider("Ukuran font legenda", 6, 24, 10)
label_font_size = st.sidebar.slider("Ukuran font label nilai", 6, 24, 9)

title_color = st.sidebar.color_picker("Warna judul", "#000000")
subtitle_color = st.sidebar.color_picker("Warna subjudul", "#595959")
axis_color = st.sidebar.color_picker("Warna label sumbu", "#000000")
tick_color = st.sidebar.color_picker("Warna tick", "#000000")
legend_color = st.sidebar.color_picker("Warna legenda", "#000000")
label_color = st.sidebar.color_picker("Warna label nilai", "#000000")

dpi = st.sidebar.selectbox("Resolusi PNG / DPI", [300, 600, 900, 1200, 1500], index=1)
transparent_background = st.sidebar.checkbox("Background PNG transparan", value=False)


# =====================================================
# MAIN DISPLAY
# =====================================================
st.info(f"Dataset aktif: **{active_source_name}** | Jumlah data setelah filter: **{len(filtered_df)} publikasi**")

metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
with metric_col1:
    st.metric("Jumlah Publikasi", len(filtered_df))
with metric_col2:
    if "Tahun" in filtered_df.columns:
        st.metric("Jumlah Tahun", filtered_df["Tahun"].nunique())
    else:
        st.metric("Jumlah Tahun", "-")
with metric_col3:
    if "Nama Dosen USU" in filtered_df.columns:
        filtered_ids = filtered_df["Publication ID"].unique()
        n_lecturers = lecturer_long[lecturer_long["Publication ID"].isin(filtered_ids)]["Nama Dosen USU (Individual)"].nunique()
        st.metric("Jumlah Dosen", n_lecturers)
    else:
        st.metric("Jumlah Dosen", "-")
with metric_col4:
    if "Terindeks Scopus" in filtered_df.columns:
        n_scopus = filtered_df[filtered_df["Terindeks Scopus"].astype(str).str.lower().isin(["ya", "yes", "true", "1"])] ["Publication ID"].nunique()
        st.metric("Terindeks Scopus", n_scopus)
    else:
        st.metric("Terindeks Scopus", "-")


tab_detail, tab_summary, tab_lecturer_year, tab_chart, tab_export = st.tabs([
    "📄 Detail Publikasi",
    "📊 Tabel Rekap",
    "👨‍🏫 Rekap Dosen per Tahun",
    "📈 Visualisasi",
    "⬇️ Download"
])

with tab_detail:
    st.subheader("Detail Data Publikasi")
    st.caption("Tabel ini mengikuti filter nama dosen, jenis publikasi, tahun, dan status Scopus yang dipilih di sidebar.")

    detail_cols = ["Publication ID"] + [col for col in display_columns if col in filtered_df.columns]
    display_dataframe_with_links(filtered_df[detail_cols], use_container_width=True)

with tab_summary:
    st.subheader("Tabel Rekap Berdasarkan Informasi yang Dipilih")

    if selected_info_table.empty:
        st.info("Pilih minimal satu kolom informasi di sidebar untuk membuat tabel rekap.")
    else:
        st.dataframe(selected_info_table, use_container_width=True, hide_index=True)

    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("#### Rekap per Dosen")
        st.dataframe(summaries["summary_by_lecturer"], use_container_width=True, hide_index=True)

        st.markdown("#### Rekap per Tahun")
        st.dataframe(summaries["summary_by_year"], use_container_width=True, hide_index=True)

    with col_b:
        st.markdown("#### Rekap Jurnal atau Prosiding")
        st.dataframe(summaries["summary_by_type"], use_container_width=True, hide_index=True)

        st.markdown("#### Rekap Terindeks Scopus")
        st.dataframe(summaries["summary_by_scopus"], use_container_width=True, hide_index=True)

    st.markdown("#### Crosstab Tahun × Jurnal/Prosiding")
    st.dataframe(summaries["cross_year_type"], use_container_width=True)

    st.markdown("#### Crosstab Tahun × Terindeks Scopus")
    st.dataframe(summaries["cross_year_scopus"], use_container_width=True)

    with st.expander("Tampilkan tabel persentase dari crosstab"):
        st.markdown("##### Persentase baris: Tahun × Jurnal/Prosiding")
        st.dataframe(format_percentage_table(summaries["cross_year_type"], basis="row"), use_container_width=True)

        st.markdown("##### Persentase baris: Tahun × Terindeks Scopus")
        st.dataframe(format_percentage_table(summaries["cross_year_scopus"], basis="row"), use_container_width=True)

with tab_lecturer_year:
    st.subheader("Ringkasan Publikasi Setiap Dosen Berdasarkan Tahun")
    st.caption(
        "Tabel ini menghitung jumlah publikasi untuk setiap dosen per tahun, "
        "dipisahkan berdasarkan jenis publikasi Jurnal atau Prosiding. "
        "Jika satu artikel memiliki beberapa dosen USU, artikel tersebut dihitung satu kali untuk masing-masing dosen terkait."
    )

    if lecturer_year_type_summary.empty:
        st.info("Data ringkasan dosen per tahun belum tersedia. Pastikan kolom Nama Dosen USU, Tahun, dan Jurnal atau Prosiding tersedia.")
    else:
        metric_a, metric_b, metric_c = st.columns(3)
        with metric_a:
            st.metric("Total Kontribusi Publikasi Dosen", int(lecturer_year_type_summary["Total Publikasi"].sum()))
        with metric_b:
            st.metric("Jumlah Dosen dalam Ringkasan", lecturer_year_type_summary["Nama Dosen USU (Individual)"].nunique())
        with metric_c:
            st.metric("Jumlah Tahun dalam Ringkasan", lecturer_year_type_summary["Tahun"].nunique())

        st.markdown("#### Tabel Dosen × Tahun × Jurnal/Prosiding")
        st.dataframe(lecturer_year_type_summary, use_container_width=True, hide_index=True)

        col_total_lecturer, col_total_year = st.columns(2)
        with col_total_lecturer:
            st.markdown("#### Total Publikasi per Dosen")
            st.dataframe(lecturer_overall_type_summary, use_container_width=True, hide_index=True)

        with col_total_year:
            st.markdown("#### Total Publikasi per Tahun")
            st.dataframe(lecturer_year_total_type_summary, use_container_width=True, hide_index=True)

        st.markdown("#### Grafik Batang Ringkasan Dosen")

        lecturer_chart_col1, lecturer_chart_col2 = st.columns(2)
        with lecturer_chart_col1:
            lecturer_year_x_options = [
                "Tahun",
                "Nama Dosen USU (Individual)",
                "Tahun | Dosen",
            ]
            lecturer_year_x_dimension = st.selectbox(
                "Sumbu X grafik ringkasan dosen",
                lecturer_year_x_options,
                index=0
            )

        with lecturer_chart_col2:
            lecturer_year_stack_options = ["Tidak ada / single bar"]
            if lecturer_year_x_dimension != "Jurnal atau Prosiding":
                lecturer_year_stack_options.append("Jurnal atau Prosiding")
            if lecturer_year_x_dimension != "Tahun":
                lecturer_year_stack_options.append("Tahun")
            if lecturer_year_x_dimension != "Nama Dosen USU (Individual)":
                lecturer_year_stack_options.append("Nama Dosen USU (Individual)")

            lecturer_year_stack_dimension = st.selectbox(
                "Stacked / legenda grafik ringkasan dosen",
                lecturer_year_stack_options,
                index=1 if "Jurnal atau Prosiding" in lecturer_year_stack_options else 0
            )

        lecturer_chart_col3, lecturer_chart_col4, lecturer_chart_col5 = st.columns(3)
        with lecturer_chart_col3:
            lecturer_year_value_mode = st.radio(
                "Nilai grafik ringkasan",
                ["Frekuensi", "Persentase (%)"],
                horizontal=True,
                key="lecturer_year_value_mode"
            )

        with lecturer_chart_col4:
            lecturer_year_percentage_basis = st.selectbox(
                "Basis persentase grafik ringkasan",
                ["Setiap bar menjadi 100%", "Dari total seluruh data"],
                index=0,
                disabled=lecturer_year_value_mode == "Frekuensi" or lecturer_year_stack_dimension == "Tidak ada / single bar",
                key="lecturer_year_percentage_basis"
            )

        with lecturer_chart_col5:
            lecturer_year_chart_type = st.radio(
                "Tipe grafik ringkasan",
                ["Stacked Bar", "Grouped Bar"],
                horizontal=True,
                disabled=lecturer_year_stack_dimension == "Tidak ada / single bar",
                key="lecturer_year_chart_type"
            )

        lecturer_year_raw_chart_table = build_lecturer_year_chart_table(
            lecturer_publication_long=lecturer_publication_long,
            x_dimension=lecturer_year_x_dimension,
            stack_dimension=lecturer_year_stack_dimension,
            value_mode=lecturer_year_value_mode,
            percentage_basis=lecturer_year_percentage_basis
        )

        lecturer_year_chart_table = reorder_axis(
            lecturer_year_raw_chart_table,
            x_order_option,
            x_custom_order,
            axis=0
        )
        if lecturer_year_stack_dimension != "Tidak ada / single bar":
            lecturer_year_chart_table = reorder_axis(
                lecturer_year_chart_table,
                stack_order_option,
                stack_custom_order,
                axis=1
            )

        if lecturer_year_chart_table.empty:
            st.warning("Tabel grafik ringkasan dosen kosong. Coba ubah pilihan sumbu X atau stacked/legenda.")
        else:
            st.markdown("##### Tabel Data Grafik Ringkasan Dosen")
            if lecturer_year_value_mode == "Persentase (%)":
                st.dataframe(lecturer_year_chart_table.map(lambda x: f"{x:.2f}%"), use_container_width=True)
            else:
                st.dataframe(lecturer_year_chart_table.astype(int), use_container_width=True)

            lecturer_year_category_colors = {}
            for i, category in enumerate(lecturer_year_chart_table.columns):
                if category == "Jumlah Publikasi":
                    continue
                lecturer_year_category_colors[category] = palette[i % len(palette)]

            lecturer_year_fig = create_publication_bar_chart(
                plot_table=lecturer_year_chart_table,
                value_mode=lecturer_year_value_mode,
                chart_type=lecturer_year_chart_type,
                category_colors=lecturer_year_category_colors,
                single_bar_color=single_bar_color,
                title="Lecturer Publication Summary by Year",
                subtitle="Journal and proceedings publication counts based on selected filters",
                x_label=lecturer_year_x_dimension,
                y_label="Frequency" if lecturer_year_value_mode == "Frekuensi" else "Percentage (%)",
                legend_title=lecturer_year_stack_dimension if lecturer_year_stack_dimension != "Tidak ada / single bar" else "",
                figure_width=figure_width,
                figure_height=figure_height,
                bar_width=bar_width,
                show_grid=show_grid,
                show_value_labels=show_value_labels,
                label_decimal_digits=label_decimal_digits,
                label_min_value=label_min_value,
                x_tick_rotation=x_tick_rotation,
                font_family=font_family,
                title_font_size=title_font_size,
                subtitle_font_size=subtitle_font_size,
                axis_font_size=axis_font_size,
                tick_font_size=tick_font_size,
                legend_font_size=legend_font_size,
                label_font_size=label_font_size,
                title_color=title_color,
                subtitle_color=subtitle_color,
                axis_color=axis_color,
                tick_color=tick_color,
                legend_color=legend_color,
                label_color=label_color,
                edge_color=edge_color,
                edge_width=edge_width,
                legend_position=legend_position
            )

            st.pyplot(lecturer_year_fig, use_container_width=False)

            lecturer_year_png_buffer = fig_to_png_bytes(
                fig=lecturer_year_fig,
                dpi=dpi,
                transparent_background=transparent_background
            )

            st.download_button(
                label=f"⬇️ Download Grafik Ringkasan Dosen PNG ({dpi} DPI)",
                data=lecturer_year_png_buffer,
                file_name="lecturer_year_publication_summary_chart.png",
                mime="image/png"
            )

            plt.close(lecturer_year_fig)

        lecturer_year_excel_output = BytesIO()
        with pd.ExcelWriter(lecturer_year_excel_output, engine="openpyxl") as writer:
            lecturer_year_type_summary.to_excel(writer, sheet_name="Dosen Tahun Jenis", index=False)
            lecturer_overall_type_summary.to_excel(writer, sheet_name="Total per Dosen", index=False)
            lecturer_year_total_type_summary.to_excel(writer, sheet_name="Total per Tahun", index=False)
            if not lecturer_publication_long.empty:
                lecturer_publication_long.to_excel(writer, sheet_name="Data Long", index=False)
            if "lecturer_year_chart_table" in locals() and not lecturer_year_chart_table.empty:
                lecturer_year_chart_table.to_excel(writer, sheet_name="Tabel Grafik")
            auto_style_workbook(writer.book)
        lecturer_year_excel_output.seek(0)

        st.download_button(
            label="⬇️ Download Tabel Ringkasan Dosen ke Excel",
            data=lecturer_year_excel_output.getvalue(),
            file_name="lecturer_year_publication_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


with tab_chart:
    st.subheader("Visualisasi Grafik Publikasi")
    st.caption("Grafik dapat diatur melalui sidebar: variabel sumbu X, stacked/legenda, warna, teks, ukuran, dan resolusi PNG.")

    st.markdown("#### Tabel Data Grafik")
    if value_mode == "Persentase (%)":
        chart_preview = chart_table.map(lambda x: f"{x:.2f}%")
    else:
        chart_preview = chart_table.astype(int)
    st.dataframe(chart_preview, use_container_width=True)

    fig = create_publication_bar_chart(
        plot_table=chart_table,
        value_mode=value_mode,
        chart_type=chart_type,
        category_colors=category_colors,
        single_bar_color=single_bar_color,
        title=chart_title,
        subtitle=chart_subtitle,
        x_label=x_label,
        y_label=y_label,
        legend_title=legend_title,
        figure_width=figure_width,
        figure_height=figure_height,
        bar_width=bar_width,
        show_grid=show_grid,
        show_value_labels=show_value_labels,
        label_decimal_digits=label_decimal_digits,
        label_min_value=label_min_value,
        x_tick_rotation=x_tick_rotation,
        font_family=font_family,
        title_font_size=title_font_size,
        subtitle_font_size=subtitle_font_size,
        axis_font_size=axis_font_size,
        tick_font_size=tick_font_size,
        legend_font_size=legend_font_size,
        label_font_size=label_font_size,
        title_color=title_color,
        subtitle_color=subtitle_color,
        axis_color=axis_color,
        tick_color=tick_color,
        legend_color=legend_color,
        label_color=label_color,
        edge_color=edge_color,
        edge_width=edge_width,
        legend_position=legend_position
    )

    st.pyplot(fig, use_container_width=False)

    png_buffer = fig_to_png_bytes(
        fig=fig,
        dpi=dpi,
        transparent_background=transparent_background
    )

    st.download_button(
        label=f"⬇️ Download Grafik PNG ({dpi} DPI)",
        data=png_buffer,
        file_name="publication_dashboard_chart.png",
        mime="image/png"
    )

    plt.close(fig)

with tab_export:
    st.subheader("Download Data dan Tabel Rekap")

    excel_bytes = dataframe_to_excel_bytes(
        filtered_df=filtered_df,
        selected_info_table=selected_info_table,
        summaries=summaries,
        chart_table=chart_table,
        lecturer_publication_long=lecturer_publication_long,
        lecturer_year_type_summary=lecturer_year_type_summary,
        lecturer_overall_type_summary=lecturer_overall_type_summary,
        lecturer_year_total_type_summary=lecturer_year_total_type_summary,
        lecturer_year_chart_table=build_lecturer_year_chart_table(
            lecturer_publication_long=lecturer_publication_long,
            x_dimension="Tahun",
            stack_dimension="Jurnal atau Prosiding",
            value_mode="Frekuensi",
            percentage_basis="Setiap bar menjadi 100%"
        ),
        selected_lecturers=selected_lecturers,
        selected_pub_types=selected_pub_types,
        selected_years=selected_years,
        selected_scopus_status=selected_scopus_status
    )

    st.download_button(
        label="⬇️ Download Dashboard Summary ke Excel",
        data=excel_bytes,
        file_name="publication_dashboard_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    csv_bytes = filtered_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="⬇️ Download Data Terfilter ke CSV",
        data=csv_bytes,
        file_name="publication_filtered_data.csv",
        mime="text/csv"
    )
